# -*- coding: utf-8 -*-
import os
import sys
import openpyxl as op
from openpyxl import Workbook


def output_files(config_params):
    path_to_output = r'D:/Code/Python_Code/Mammo/src/out_res/run/'

    # check output_folder path
    if not os.path.exists(path_to_output):
        print(
            "Error! config file path does not exist! This code needs the same path to store the output files and model.")
        sys.exit()

    if config_params['randseeddata'] != config_params['randseedother']:
        rand_seed = str(config_params['randseedother']) + '_' + str(config_params['randseeddata'])
    else:
        rand_seed = str(config_params['randseeddata'])

    path_to_model = path_to_output + "model_" + str(rand_seed) + ".tar"
    path_to_results = path_to_output + "result_" + str(rand_seed) + ".xlsx"
    path_to_results_text = path_to_output + "result_" + str(rand_seed) + ".txt"
    path_to_learning_curve = path_to_output + "learningcurve_" + str(rand_seed) + ".png"
    path_to_log_file = path_to_output + "log_" + str(rand_seed) + ".txt"

    # set file path
    if os.path.isfile(path_to_results):
        wb = op.load_workbook(path_to_results)
        sheet1 = wb['train_val_results']
        sheet2 = wb['test_results']
    else:
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "train_val_results"
        header = ['Epoch', 'lr', 'Avg Loss Train', 'Accuracy Train', 'F1macro Train', 'Recall Train', 'Speci Train',
                  'Avg Loss Val', 'Accuracy Val', 'F1macro Val', 'Recall Val', 'Speci Val', 'AUC Val']

        sheet1.append(header)
        sheet2 = wb.create_sheet('test_results')
        sheet2.append(
            ['Loss', 'PrecisionBin', 'PrecisionMicro', 'PrecisionMacro', 'RecallBin', 'RecallMicro', 'RecallMacro',
             'F1Bin', 'F1Micro', 'F1macro', 'F1wtmacro', 'Acc', 'Cohens Kappa', 'AUC', 'epoch'])

    wb.save(path_to_results)

    return path_to_model, path_to_results, path_to_results_text, path_to_learning_curve, path_to_log_file
